import pandas as pd
import openpyxl

import glob
tracker_file = r'E:\MBF Rollout Dashboard\On Air Progress Tracker.xlsx'
xlsm_files = glob.glob(r'E:\MBF Rollout Dashboard\60086951_56A0US7_*.xlsm')
if xlsm_files:
    xlsm_file = sorted(xlsm_files)[-1]
else:
    xlsm_file = r'E:\MBF Rollout Dashboard\60086951_56A0US7_20260518214245.xlsm'

print(f"Loading ISDP data from: {xlsm_file}...")
df_isdp = pd.read_excel(xlsm_file, sheet_name='Site Rollout Plan', header=[0, 1])

site_data = {}
for idx, row in df_isdp.iterrows():
    site_name = row.get(('Customer Site Name', 'Unnamed: 1_level_1'))
    if pd.isna(site_name):
        continue
        
    on_air_4g = row.get(('On-Air 4G', 'Actual End Date'))
    on_air_5g = row.get(('On-Air 5G', 'Actual End Date'))
    physical_on_air = row.get(('Physical Site On Air', 'Actual End Date'))
    rfi = row.get(('Ready For Installation', 'Actual End Date'))
    
    site_data[str(site_name).strip()] = {
        '4g': on_air_4g if pd.notna(on_air_4g) else None,
        '5g': on_air_5g if pd.notna(on_air_5g) else None,
        'physical': physical_on_air if pd.notna(physical_on_air) else None,
        'rfi': rfi if pd.notna(rfi) else None
    }

print(f"Loaded {len(site_data)} sites from ISDP")

wb = openpyxl.load_workbook(tracker_file)
sheets = [
    'Ha Noi Progress', 
    'North 5G Progress', 
    'Middle Swap Progress ', 
    'Middle 5G Progress', 
    'South Swap Progress ', 
    'South 5G  Progress'
]

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    
    # find header row (typically row 2)
    header_row_idx = 2
    
    # build column map
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row_idx, column=col_idx).value
        top_val = ws.cell(row=1, column=col_idx).value
        
        if cell_val:
            val_str = str(cell_val).strip()
            
            if val_str == 'Site Name':
                col_map['Site Name'] = col_idx
            elif val_str == 'Site Status':
                col_map['Site Status'] = col_idx
            elif val_str == 'On-air Day':
                # check top_val to determine which band
                # it could be merged, so we need to track the last seen top_val if it's None
                band = "unknown"
                # scan backwards for merged cell value if top_val is None
                for c in range(col_idx, 0, -1):
                    t_val = ws.cell(row=1, column=c).value
                    if t_val:
                        band = str(t_val).strip()
                        break
                
                if '4G' in band:
                    col_map['On-air Day 4G'] = col_idx
                elif '5G' in band or '3.8G' in band or '2.6G' in band:
                    if 'On-air Day 5G_1' not in col_map:
                        col_map['On-air Day 5G_1'] = col_idx
                    else:
                        col_map['On-air Day 5G_2'] = col_idx
            elif val_str == 'RFI Date':
                col_map['RFI Date'] = col_idx

    # If RFI Date doesn't exist, append it
    if 'RFI Date' not in col_map:
        new_col = ws.max_column + 1
        ws.cell(row=header_row_idx, column=new_col).value = 'RFI Date'
        col_map['RFI Date'] = new_col

    print(f"Sheet {sheet_name} col map: {col_map}")

    # update rows
    updated = 0
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        site_name_cell = ws.cell(row=row_idx, column=col_map.get('Site Name', 0))
        if site_name_cell and site_name_cell.value:
            s_name = str(site_name_cell.value).strip()
            if s_name in site_data:
                data = site_data[s_name]
                
                # Update On-air Day 4G
                if 'On-air Day 4G' in col_map and data['4g']:
                    ws.cell(row=row_idx, column=col_map['On-air Day 4G']).value = data['4g']
                    ws.cell(row=row_idx, column=col_map['Site Status']).value = 'On Air'
                
                # Update On-air Day 5G
                if 'On-air Day 5G_1' in col_map and data['5g']:
                    ws.cell(row=row_idx, column=col_map['On-air Day 5G_1']).value = data['5g']
                    ws.cell(row=row_idx, column=col_map['Site Status']).value = 'On Air'
                
                if 'On-air Day 5G_2' in col_map and data['5g']:
                    ws.cell(row=row_idx, column=col_map['On-air Day 5G_2']).value = data['5g']
                    ws.cell(row=row_idx, column=col_map['Site Status']).value = 'On Air'
                    
                # Update RFI
                if 'RFI Date' in col_map and data['rfi']:
                    ws.cell(row=row_idx, column=col_map['RFI Date']).value = data['rfi']
                    
                updated += 1

    print(f"Updated {updated} rows in {sheet_name}")

wb.save(tracker_file)
print("Saved tracker file successfully.")
